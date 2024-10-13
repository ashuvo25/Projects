import os
import pandas as pd
from datetime import datetime
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment
from openpyxl.utils import get_column_letter
from tkinter import Tk, filedialog, Label, Button, messagebox

# Function to get file details
def get_files_details(folder_path):
    file_data = []
    for root, dirs, files in os.walk(folder_path):
        for file in files:
            file_path = os.path.join(root, file)
            file_size = os.path.getsize(file_path)
            # Convert file size to appropriate units
            if file_size < 1024:  # Less than 1 KB
                file_size_display = f"{file_size} Bytes"
            elif file_size < 1024 ** 2:  # Less than 1 MB
                file_size_display = f"{file_size / 1024:.2f} KB"
            elif file_size < 1024 ** 3:  # Less than 1 GB
                file_size_display = f"{file_size / (1024 ** 2):.2f} MB"
            else:  # Greater than or equal to 1 GB
                file_size_display = f"{file_size / (1024 ** 3):.2f} GB"
            
            file_data.append((file, file_size_display, datetime.now().date())) 
    return file_data

# Function to check if file is already in the Excel sheet
def is_file_in_sheet(worksheet, file_name):
    for row in worksheet.iter_rows(min_row=2, values_only=True):
        if row[0] == file_name:
            return True
    return False

# Function to write data to Excel
def write_to_excel(folder_path, excel_name='files_data.xlsx'):
    file_details = get_files_details(folder_path)

    code_folder = os.path.dirname(os.path.abspath(__file__))
    excel_path = os.path.join(code_folder, excel_name)

    # Get the last folder name from the folder path
    last_folder_name = os.path.basename(os.path.normpath(folder_path))

    if os.path.exists(excel_path):
        wb = load_workbook(excel_path)
    else:
        wb = Workbook()

    # Create or use the existing sheet for the last folder name
    if last_folder_name not in wb.sheetnames:
        ws = wb.create_sheet(title=last_folder_name)
        ws.append(['File Name', 'File Size', 'Date Stored'])  
    else:
        ws = wb[last_folder_name]

    # Append only new file details
    new_files_count = 0
    for detail in file_details:
        file_name, file_size, date_stored = detail
        if not is_file_in_sheet(ws, file_name):
            ws.append([file_name, file_size, date_stored])
            new_files_count += 1

    for sheet in wb.sheetnames:
        adjust_excel_formatting(wb[sheet])

    wb.save(excel_path)
    
    if new_files_count > 0:
        messagebox.showinfo("Success", f"{new_files_count} new files added to '{excel_name}'.")
    else:
        messagebox.showinfo("Info", "No new files to add.")

# Function to adjust Excel formatting
def adjust_excel_formatting(worksheet):
    for row in worksheet.iter_rows(min_row=2):
        worksheet.row_dimensions[row[0].row].height = 30
    for col in range(1, worksheet.max_column + 1):
        worksheet.column_dimensions[get_column_letter(col)].width = 25

# GUI Application
def open_folder():
    folder_path = filedialog.askdirectory()
    if folder_path:
        write_to_excel(folder_path)

def create_gui():
    root = Tk()
    root.title("File Details Collector")
    root.geometry("400x200")

    label = Label(root, text="Click below to select a folder", font=("Arial", 14))
    label.pack(pady=20)

    browse_button = Button(root, text="Browse Folder", command=open_folder, font=("Arial", 12))
    browse_button.pack(pady=10)

    exit_button = Button(root, text="Exit", command=root.quit, font=("Arial", 12))
    exit_button.pack(pady=10)

    root.mainloop()

# Main function to run the GUI
if __name__ == "__main__":
    create_gui()
