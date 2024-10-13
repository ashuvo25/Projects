import os
import pandas as pd
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Alignment
from openpyxl.utils import get_column_letter

# Function to get file details
def get_files_details(folder_path):
    file_data = []
    for root, dirs, files in os.walk(folder_path):
        for file in files:
            file_path = os.path.join(root, file)
            file_size_mb = os.path.getsize(file_path) / (1024 * 1024) 
            file_data.append((file, file_size_mb, datetime.now()))
    return file_data

# Function to write data to Excel
def write_to_excel(folder_path, excel_name='files_data.xlsx'):
    company_name = os.path.basename(os.path.normpath(folder_path))
    file_details = get_files_details(folder_path)
    df = pd.DataFrame(file_details, columns=['File Name', 'File Size (MB)', 'Date Stored'])
    df['Company Name'] = company_name
    df = df[['Company Name', 'File Name', 'File Size (MB)', 'Date Stored']]

    df = df.sort_values(by='Company Name')

    code_folder = os.path.dirname(os.path.abspath(__file__))
    excel_path = os.path.join(code_folder, excel_name)

    try:
        existing_df = pd.read_excel(excel_path)
        combined_df = pd.concat([existing_df, df], ignore_index=True)
        combined_df = combined_df.sort_values(by='Company Name')
        combined_df.to_excel(excel_path, index=False)
    except FileNotFoundError:
        df.to_excel(excel_path, index=False)

    adjust_excel_formatting(excel_path)

def adjust_excel_formatting(excel_path):
    wb = load_workbook(excel_path)
    ws = wb.active
 
    for row in ws.iter_rows(min_row=2):
        ws.row_dimensions[row[0].row].height = 30  
    
    for col in range(1, ws.max_column + 1):
        ws.column_dimensions[get_column_letter(col)].width = 20  
    merge_cells_with_same_value(ws, 'A')
    wb.save(excel_path)

def merge_cells_with_same_value(worksheet, column_letter):
    start_row = None
    last_value = None

    for row in range(2, worksheet.max_row + 1):
        current_value = worksheet[f'{column_letter}{row}'].value

        if current_value != last_value:
            if start_row is not None and row - start_row > 1:
                worksheet.merge_cells(f'{column_letter}{start_row}:{column_letter}{row - 1}')
                worksheet[f'{column_letter}{start_row}'].alignment = Alignment(vertical='center')

            start_row = row
            last_value = current_value
    if start_row is not None and worksheet.max_row - start_row > 0:
        worksheet.merge_cells(f'{column_letter}{start_row}:{column_letter}{worksheet.max_row}')
        worksheet[f'{column_letter}{start_row}'].alignment = Alignment(vertical='center')

# Main function to take input and execute
def main():
    folder_path = input("Enter the folder path: ") 
    write_to_excel(folder_path)
    print("File details have been stored in 'files_data.xlsx' in the code folder with formatting adjustments.")

if __name__ == "__main__":
    main()
